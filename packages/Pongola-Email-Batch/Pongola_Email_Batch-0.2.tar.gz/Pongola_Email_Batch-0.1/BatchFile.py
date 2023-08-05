from __future__ import print_function
from openpyxl import load_workbook
import requests
import sys
import re
from time import sleep
import datetime
import os.path
import smtplib
from email.MIMEMultipart import MIMEMultipart
from email.MIMEText import MIMEText
from email.MIMEBase import MIMEBase
from email import encoders


#accessing the excell file
wb = load_workbook("EmailList.xlsx", data_only=True)
shF = wb["Filters"]
shG = wb['Growers']

log =[]
base_url = "http://www.emailList.com"
#get the total number of columns in the excell sheet
row_count = shG.max_row
col_count = shG.max_column
col_countF = shF.max_column
dictFilters ={}
growerString="GrowerCode"

class ProgressBar(object):
    DEFAULT = 'Progress: %(bar)s %(percent)3d%%'
    FULL = '%(bar)s %(current)d/%(total)d (%(percent)3d%%) %(remaining)d to go'

    def __init__(self, total, width=40, fmt=DEFAULT, symbol='#',
                 output=sys.stderr):
        assert len(symbol) == 1

        self.total = total
        self.width = width
        self.symbol = symbol
        self.output = output
        self.fmt = re.sub(r'(?P<name>%\(.+?\))d',
            r'\g<name>%dd' % len(str(total)), fmt)

        self.current = 0

    def __call__(self):
        percent = self.current / float(self.total)
        size = int(self.width * percent)
        remaining = self.total - self.current
        bar = '[' + self.symbol * size + ' ' * (self.width - size) +']'

        args = {
            'total': self.total,
            'bar': bar,
            'current': self.current,
            'percent': percent * 100,
            'remaining': remaining
        }
        print('\r'+self.fmt % args, end=' ')


    def done(self):
        self.current = self.total
        self()


total_Columns=col_countF+1
message="Email Sent To: "
for col in range(3,total_Columns):
    Keys=shF.cell(1, col).value
    for row in range(1,row_count):
        value=shF.cell(2, col).value
    dictFilters[Keys] = value

#create dictionary to extract Growers info sheet
myDict = {}

#getting Testing_mode value

Testing_mode = shF["A2"].value
row_count = row_count+1

path=os.path.join(os.path.expanduser('~'),'PycharmProjects','pongolaEmails')
now = datetime.datetime.now()
folder_name = now.strftime("%Y_%m_%d-%H%M%S")
os.mkdir(folder_name)
path=os.path.join(os.path.expanduser('~'),'PycharmProjects','pongolaEmails',folder_name)

if not os.path.exists(path):
    os.makedirs(path)
fname='_'.join([now.strftime("%Y_%m_%d"),'.log'])
log_file_path= os.path.join(path,fname)
DTS= now.strftime("%Y_%m_%d-%H%M%S")
for i in range(2,row_count) :
    if Testing_mode:
        email = shG['D' + str(i)].value
        name = shG['B'+str(i)].value
        Grower_code = shG['A'+str(i)].value
        Email_Address = email
        myDict[Email_Address] = Grower_code , name
        for keys in dictFilters.items():
            URL_list = [base_url,':'.join([str(item) for item in keys if item is not None]),':'.join([growerString,str(Grower_code)])]
            URL = '/'.join(URL_list)

            print(URL)
            with open(log_file_path, 'a') as lf:
                lf.write('\n'+'   '.join([DTS,URL]))
            lf.close()

            r = requests.get(URL, auth=('user', 'pass'))
            r.status_code
            print(r)
            with open(log_file_path, 'a') as lf:
                lf.write('\n'+'   '.join([DTS,str(r)]))
            lf.close()
            now = datetime.datetime.now()
            folder_name=os.path.join(path,Grower_code)
            print("Making directory " + folder_name)
            with open(log_file_path, 'a') as lf:
                lf.write('\n'+'   '.join([DTS,folder_name]))
            lf.close()
            r.text
            os.mkdir(folder_name)
            extension=".html"
            name_of_file=''.join(['_'.join([str(Grower_code),'GrowerReport',now.strftime("%Y_%m_%d-%H%M%S")]),extension])
            file_path = os.path.join(path,folder_name,name_of_file)
            try:
                with open(file_path, 'w') as f:
                        f.write(r.text)
            except:
                with open(file_path, 'w') as f:
                    f.write(r.text.encode('UTF-8'))
#****************************EMAIL****************************************************************************
            fromaddr = "obhero3216@gmail.com"
            toaddr = '4c5q6xtr6@educationgroup.co.za'

            msg = MIMEMultipart()

            msg['From'] = fromaddr
            msg['To'] = toaddr
            msg['Subject'] = ':'.join([str(name),str(Grower_code)])

            body = "Dear "+name+"\r\r\rPlease find attached summary report for " +Grower_code+"\r\r\r\rKind regards,\r\r'--'\rPongola GIS Team"



            msg.attach(MIMEText(body, 'plain'))

            filename = name_of_file
            attachment = open(file_path, "rb")

            part = MIMEBase('application', 'octet-stream')
            part.set_payload((attachment).read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', "attachment; filename= %s" % filename)

            msg.attach(part)

            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login(fromaddr, "66592632")
            text = msg.as_string()
            server.sendmail(fromaddr, toaddr, text)
            server.quit()


        progress = ProgressBar(100, fmt=ProgressBar.FULL)
        for x in range(100):
            progress.current += 1
            progress()
            sleep(0.1)
        progress.done()
        print('\r'+'**'+message+Email_Address+'**')
        with open(log_file_path, 'a') as lf:
                lf.write('\n'+'   '.join([DTS,message])+' '.join([message,email]))
        lf.close()
        i=i+1


    else:
        email = shG['C' + str(i)].value
        name = shG['B'+str(i)].value
        Grower_code = shG['A'+str(i)].value
        Email_Address = email
        myDict[Email_Address] = Grower_code , name
        for keys in dictFilters.items():
            URL_list = [base_url,':'.join([str(item) for item in keys if item is not None]),':'.join([growerString,str(Grower_code)])]
            URL = '/'.join(URL_list)

            print(URL)
            with open(log_file_path, 'a') as lf:
                lf.write('\n'+'   '.join([DTS,URL]))
            lf.close()

            r = requests.get(URL, auth=('user', 'pass'))
            r.status_code
            print(r)
            with open(log_file_path, 'a') as lf:
                lf.write('\n'+'   '.join([DTS,str(r)]))
            lf.close()
            now = datetime.datetime.now()
            folder_name=os.path.join(path,Grower_code)
            print("Making directory " + folder_name)
            with open(log_file_path, 'a') as lf:
                lf.write('\n'+'   '.join([DTS,folder_name]))
            lf.close()
            r.text
            os.mkdir(folder_name)
            extension=".html"
            name_of_file=''.join(['_'.join([str(Grower_code),'GrowerReport',now.strftime("%Y_%m_%d-%H%M%S")]),extension])
            file_path = os.path.join(path,folder_name,name_of_file)
            try:
                with open(file_path, 'w') as f:
                        f.write(r.text)
            except:
                with open(file_path, 'w') as f:
                    f.write(r.text.encode('UTF-8'))
#****************************EMAIL****************************************************************************
            fromaddr = "obhero3216@gmail.com"
            toaddr = Email_Address

            msg = MIMEMultipart()

            msg['From'] = fromaddr
            msg['To'] = toaddr
            msg['Subject'] = ':'.join([str(name),str(Grower_code)])

            body = "Dear Sir/Mam. \r\r\rplease do find the attached document.\r\r\r\rRegards:\rPongola GIS"




            msg.attach(MIMEText(body, 'plain'))

            filename = name_of_file
            attachment = open(file_path, "rb")

            part = MIMEBase('application', 'octet-stream')
            part.set_payload((attachment).read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', "attachment; filename= %s" % filename)

            msg.attach(part)

            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login(fromaddr, "66592632")
            text = msg.as_string()
            server.sendmail(fromaddr, toaddr, text)
            server.quit()


        progress = ProgressBar(100, fmt=ProgressBar.FULL)
        for x in range(100):
            progress.current += 1
            progress()
            sleep(0.1)
        progress.done()
        print('\r'+'**'+message+Email_Address+'**')
        i=i+1
        with open(log_file_path, 'a') as lf:
                lf.write('\n'+'   '.join([DTS,message])+' '.join([message,email]))
        lf.close()
