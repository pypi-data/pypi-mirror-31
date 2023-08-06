# -*- coding: utf-8 -*-
import sys
from antlr4 import *
from AntlrLexer import AntlrLexer
from AntlrParser import AntlrParser
from AntlrListener import AntlrListener

import zipfile
import os
import codecs
import xlwt
# workbook相关
from openpyxl.workbook import Workbook
# ExcelWriter，封装了很强大的excel写的功能
from openpyxl.writer.excel import ExcelWriter
# 一个eggache的数字转为列字母的方法
from openpyxl.utils import get_column_letter
from openpyxl.reader.excel import load_workbook
import time
# from numba import jit
from timeit import timeit
import json
import re
import datetime
import string
from zipfile import ZipFile, ZIP_DEFLATED



class HandleExcel():
    def __init__(self):
        self.head_row_labels = [u'Name', u'Abstraction', u'Parallelism', u'LogicalThinking', u'Synchronization',
                                u'FlowControl', u'UserInteractivity', u'DataRepresentation', u'CodeOrganization', u'Content']

    def read_from_file(self):
        score = {}
        path = os.path.abspath('.')
        filepath = path + '/test'
        pathDir = os.listdir(filepath)
        for allDir in pathDir:
            child = os.path.join(filepath, allDir)
            print("name=", allDir)
            listener_score = gen(child)[0]
            score[allDir] = listener_score
        return score

    def write_to_excel_with_openpyxl(self, records, head_row, save_excel_name):
        # 新建一个workbook
        wb = Workbook()
        archive = ZipFile(save_excel_name, 'w', ZIP_DEFLATED)
        ew = ExcelWriter(workbook=wb, archive=archive)
        # 新建一个excelWriter
        # ew = ExcelWriter(workbook=wb, archive=archive)
        # 设置文件输出路径与名称
        dest_filename = save_excel_name
        # 第一个sheet是ws
        ws = wb.worksheets[0]
        # 设置ws的名称
        ws.title = "range names"
        # 写第一行，标题行
        for h_x in range(1, len(head_row) + 1):
            h_col = get_column_letter(h_x)
            ws.cell('%s%s' % (h_col, 1)).value = '%s' % (head_row[h_x - 1])
        # 写第二行及其以后的那些行
        row = 2
        for name in records:
            ws.cell('%s%s' % ('A', row)).value = '%s' % name
            col = 2
            for point in records[name]:
                col_num = get_column_letter(col)
                ws.cell('%s%s' % (col_num, row)).value = '%s' % records[name][point]
                col += 1
            row += 1
        ew.save(filename=dest_filename)

    def run_main_save_to_excel_with_openpyxl(self):
        dataset_list = self.read_from_file()
        '''test use openpyxl to handle EXCEL 2007'''
        head_row_label = self.head_row_labels
        save_name = "test_openpyxl.xlsx"
        self.write_to_excel_with_openpyxl(dataset_list, head_row_label, save_name)

def unzip_scratch(filename):
    """
    unzip scratch project and extract project.json file   
    :param filename: filename fo scratch project 
    :return: null or project.json content
    """
    zfile = zipfile.ZipFile(filename, 'r')
    if "project.json" in zfile.namelist():
        data = zfile.read("project.json")

        pattern1 = re.compile(r'"objName".+?,')
        pattern2 = re.compile(r'"scripts".+?(?= "sounds"| "costumes": \[\{)')
        encoded_data = codecs.decode(data, 'utf-8', 'strict')
        useful_data1 = pattern1.findall(encoded_data, re.S)
        useful_data2 = pattern2.findall(encoded_data, re.S)
        useful_data = ''.join(useful_data1 + useful_data2)
        useful_data = "{" + useful_data[:-1] + "}"
        # print("data_after", useful_data)
        return useful_data
    else:
        return None

def gen(argv):
    raw_json = unzip_scratch(argv)
    # encoded_json = codecs.decode(raw_json, 'utf-8', 'strict')
    input = InputStream(raw_json)
    if not input:
        return
    lexer = AntlrLexer(input)
    stream = CommonTokenStream(lexer)
    parser = AntlrParser(stream)
    tree = parser.json()
    walker = ParseTreeWalker()
    listener = AntlrListener()
    walker.walk(listener, tree)
    return listener.score, listener.hint, listener.profile
#
if __name__ == '__main__':
	gen(sys.argv[1])