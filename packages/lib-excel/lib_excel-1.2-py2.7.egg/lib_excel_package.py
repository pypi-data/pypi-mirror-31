

import openpyxl
import lib_common


#################################################################



def create_workbook():
    wb = openpyxl.Workbook()
    return wb

def open_workbook(pathname):
    wb = openpyxl.load_workbook(pathname)
    return wb

def get_sheet_name_list(wb):
    ws_name_list = wb.get_sheet_names()
    return lib_common.convert_to_string(ws_name_list)

def save_workbook(wb, pathname):
    file_close = False
    while True:
        try:
            wb.save(pathname)
            if file_close: print 'Continuing ....'
            break
        except:
            file_close = True
            print 'The following file must be closed: %s' % pathname
            prompt = 'Press [Enter] when ready to proceed...'
            lib_common.prompt_processing(prompt)

def get_worksheet(wb, ws_title):
    ws = wb[ws_title]   # getting a worksheet object from the tab name
    return ws

def get_worksheet_2(wb, ws_title):
    ws = wb.get_sheet_by_name(ws_title)    # Same thing (?)
    return ws

def get_max_row_on_worksheet(wb, ws):
    return ws.max_row

def get_max_column_on_worksheet(wb, ws):
    return ws.max_column

def create_worksheet(wb, ws_title):
    ws = wb.create_sheet(title=ws_title)
    return ws

def get_active_worksheet(wb):
    ws = wb.active
    return ws

def get_active_worksheet_name(wb):
    ws_title =  wb.active.title
    return ws_title

def set_new_worksheet_name(wb, ws, ws_title):
    ws.title = ws_title     # New tab name

def column_letter_to_number(col_letter):
    col_number = openpyxl.cell.column_index_from_string(col_letter)  # column string ==> number
    return col_number

def column_number_to_letter(col_number):
    col_letter = openpyxl.cell.get_column_letter(col_number)         # column number ==> string
    return col_letter

def get_coordinate_from_string(cell_location_string, col_type='letter'):
    column, row = openpyxl.utils.coordinate_from_string(cell_location_string)
    if col_type != 'letter':
        column = column_letter_to_number(column)
    return column, row

def get_cell_value(wb, ws, row, column):
    value = ws.cell(row=row, column=column).value             # cell Value
    return value

def set_cell_value(wb, ws, row, column, value):
    ws.cell(row=row, column=column).value = value            # cell Value

def get_color(color_id):
    # color_id = GREEN, WHITE, YELLOW, DARKYELLOW, RED
    color = openpyxl.styles.colors.color_id
    return color

def get_fill_color(color_id):
    if color_id == 'GREEN':
        color_fill = openpyxl.styles.PatternFill(start_color=openpyxl.styles.colors.GREEN, end_color=openpyxl.styles.colors.GREEN, fill_type='solid')
    elif color_id == 'YELLOW':
        color_fill = openpyxl.styles.PatternFill(start_color=openpyxl.styles.colors.YELLOW, end_color=openpyxl.styles.colors.YELLOW, fill_type='solid')
    elif color_id == 'DARKYELLOW':
        color_fill = openpyxl.styles.PatternFill(start_color=openpyxl.styles.colors.DARKYELLOW, end_color=openpyxl.styles.colors.DARKYELLOW, fill_type='solid')
    elif color_id == 'RED':
        color_fill = openpyxl.styles.PatternFill(start_color=openpyxl.styles.colors.RED, end_color=openpyxl.styles.colors.RED, fill_type='solid')
    else:   # WHITE
        color_fill = openpyxl.styles.PatternFill(start_color=openpyxl.styles.colors.WHITE, end_color=openpyxl.styles.colors.WHITE, fill_type='solid')

    return color_fill

def get_font_color(color_id, bold=False):
     if color_id == 'GREEN':
        color_font = openpyxl.styles.Font(bold=bold, color=openpyxl.styles.colors.GREEN)
     elif color_id == 'YELLOW':
        color_font = openpyxl.styles.Font(bold=bold, color=openpyxl.styles.colors.YELLOW)
     elif color_id == 'DARKYELLOW':
        color_font = openpyxl.styles.Font(bold=bold, color=openpyxl.styles.colors.DARKYELLOW)
     elif color_id == 'RED':
        color_font = openpyxl.styles.Font(bold=bold, color=openpyxl.styles.colors.RED)
     elif color_id == 'WHITE':
        color_font = openpyxl.styles.Font(bold=bold, color=openpyxl.styles.colors.WHITE)
     else:   # BLACK
        color_font = openpyxl.styles.Font(bold=bold, color=openpyxl.styles.colors.BLACK)

     return color_font


#   cell area = top_lefr=(col_start, row_start) area_dimension=(col_dim, row_dim)
#
#   sides = {left | right | top | bottom: thin | thick | none}
#   edges = {left | right | top | bottom: thin | thick | none}
def set_border(ws, col_start, row_start, col_dim, row_dim, sides={}, edges={}):
    for n_row in range(row_start, row_start + row_dim):
        for n_col in range(col_start, col_start + col_dim):
            border = openpyxl.styles.Border(
                left=ws.cell(column=n_col, row=n_row).border.left,
                right=ws.cell(column=n_col, row=n_row).border.right,
                top=ws.cell(column=n_col, row=n_row).border.top,
                bottom=ws.cell(column=n_col, row=n_row).border.bottom
            )
            if sides != {}:
                if sides['left'] != 'none':   border.left   = openpyxl.styles.Side(border_style=sides['left'], color="FF000000")
                if sides['right'] != 'none':  border.right  = openpyxl.styles.Side(border_style=sides['right'], color="FF000000")
                if sides['top'] != 'none':    border.top    = openpyxl.styles.Side(border_style=sides['top'], color="FF000000")
                if sides['bottom'] != 'none': border.bottom = openpyxl.styles.Side(border_style=sides['bottom'], color="FF000000")
#                ws.cell(column=n_col, row=n_row).border = border

            if edges != {}:
                if n_row == row_start:
                    if edges['top'] != 'none':
                        border.top = openpyxl.styles.Side(border_style=edges['top'], color="FF000000")
                if n_row == row_start + row_dim - 1:
                    if edges['bottom'] != 'none':
                        border.bottom = openpyxl.styles.Side(border_style=edges['bottom'], color="FF000000")
                if n_col == col_start:
                    if edges['left'] != 'none':
                        border.left = openpyxl.styles.Side(border_style=edges['left'], color="FF000000")
                if n_col == col_start + col_dim - 1:
                    if edges['right'] != 'none':
                        border.right = openpyxl.styles.Side(border_style=edges['right'], color="FF000000")

            ws.cell(column=n_col, row=n_row).border = border


def cell_formattingX(operator, formula, color_fill, color_font):
    # fill = color_fill = get_fill_color(color_id)
    # formula = ['"DN"']
    cell_format = openpyxl.formatting.rule.CellIsRule(operator='equal', formula=formula, fill=color_fill, font=color_font)
    return cell_format

def cell_formatting(operator, formula, color_fill='WHITE', color_font='BLACK', bold=False):
    # operator = 'equal'
    # formula = ['"DN"']
    color_fill_obj = get_fill_color(color_fill)
    color_font_obj = get_font_color(color_font, bold)
    cell_format = openpyxl.formatting.rule.CellIsRule(operator='equal', formula=formula, fill=color_fill_obj, font=color_font_obj)
    return cell_format

def cell_area_conditional_formatting(wb, ws, cell_area, cell_format):
    ws.conditional_formatting.add(cell_area, cell_format)

